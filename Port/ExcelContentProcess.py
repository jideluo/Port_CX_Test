
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import  datetime
import os,sys
from Mysql import insertDb
def  ExcelMean(Excel):#求平均值
    # Excel = openpyxl.load_workbook("PostTest.xlsx")
    sheet = Excel.get_sheet_by_name('post')
    maxRow = sheet.max_row
    maxColum = sheet.max_column
    print(maxColum,maxRow)
    for x in range(3, maxRow+1):
        Mean =0
        for y in range(7, maxColum):
            Mean+= sheet.cell(row=x, column=y).value
        sheet.cell(row=x, column=12).value ="{:.3f}".format(Mean/5)#将平均值取小数点三位

def ExcelClear(Excel):#清除测试结果
    # Excel = openpyxl.load_workbook("PostTest.xlsx")  # 打开excel文件
    sheet = Excel.get_sheet_by_name('post')
    maxRow = sheet.max_row
    maxColum = sheet.max_column
    print(maxRow,maxColum)
    for x in range(3, maxRow+1):
        for y in range(5, maxColum+1):
            sheet.cell(row=x, column=y).value = ""

def ExcleBorder(Excel):#设置边框
    sheet = Excel.get_sheet_by_name('首页')
    maxRow = sheet.max_row
    maxColum = sheet.max_column
    # sheet.print_area='A1:F10'
    # border = Border(left=Side(style='medium',color='FF000000'),right=Side(style='medium',color='FF000000'),top=Side(style='medium',color='FF000000'),bottom=Side(style='medium',color='FF000000'),diagonal=Side(style='medium',color='FF000000'),diagonal_direction=0,outline=Side(style='medium',color='FF000000'),vertical=Side(style='medium',color='FF000000'),horizontal=Side(style='medium',color='FF000000'))
    #border =Border(left=Side(border_style=None,color = 'FF000000'),right = Side(border_style=None,color = 'FF000000'),top = Side(border_style=None,color = 'FF000000'),bottom = Side(border_style=None,color = 'FF000000'),diagonal = Side(border_style=None,color = 'FF000000'),diagonal_direction = 0,outline = Side(border_style=None,color = 'FF000000'),vertical = Side(border_style=None,color = 'FF000000'),horizontal = Side(border_style=None,color = 'FF000000'))
    border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                    top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'),
                    diagonal=Side(style='thin', color='FF000000'), diagonal_direction=0,
                    outline=Side(style='thin', color='FF000000'), vertical=Side(style='thin', color='FF000000'),
                    horizontal=Side(style='thin', color='FF000000'))
    for x in range(1,maxColum-1):
        for y in range(1, maxRow-1):
            sheet.cell(row=x, column=y).border=border#设置区域边框

def ExcleTestResult(Excel):#填写测试结果
    t1 = open('TestResult.txt', 'w')
    sheet = Excel.get_sheet_by_name("post")
    maxColum = sheet.max_column
    ResultPass=0
    ResultFail=0
    data={ }#存储数据
    # p =Person()
    for x in range(3, maxColum +1):
        Result =sheet.cell(row=x, column=5).value
        functionName = sheet.cell(row=x, column=2).value
        averageData  = sheet.cell(row=x, column=12).value
        netstat =sheet.cell(row=x, column=6).value
        data[functionName]=averageData,netstat#将数据存储到字典中
        if(Result is 'Y'):
            ResultPass = ResultPass + 1
        else:
            ResultFail = ResultFail + 1
    print(data)
    insertDb(data)#提交数据到数据库
    sheet1 = Excel.get_sheet_by_name("首页")
    if(ResultFail==0):
        sheet1.cell(row=7, column=2).value="PASS"
        t1.write("0")
        t1.close()
    else:
        sheet1.cell(row=7, column=2).value = "FAIL"
        t1.write("1")
        t1.close()

def TestPersonnel(Excel):#填写测试信息
    sheet = Excel.get_sheet_by_name("首页")
    sheet.cell(row=3, column=2).value="Automaton"
    sheet.cell(row=3, column=5).value=datetime.datetime.now()

