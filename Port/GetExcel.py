
# 获取Excel表数据
# 创建者:Leon
# 创建日期：2017-08-22 11:04:13
# -*- coding: utf-8 -*-
import  openpyxl
from Post import post,postBody
from ExcelContentProcess import ExcelClear,ExcleBorder,ExcelMean,ExcleTestResult,TestPersonnel
# from ExcelContentProcess import ExcleClear
def OpenExcel(servicer_url,Excel_File,Operation_Number):

    Excel = openpyxl .load_workbook("PostTest.xlsx")#打开excel文件
    ExcelClear(Excel)
    sheet = Excel.get_sheet_by_name('post')
    maxColum  = sheet.max_column
    maxRow = sheet.max_row
    for number in range(Operation_Number):
        for i in range(3, maxRow+1):
            post = sheet.cell(row=i, column=3).value
            value = sheet.cell(row=i, column=4).value
            try:
                value = eval(value)#str转字典时存在反斜杠转码问题，需要解决
            except Exception as err:
                print(err)

            url = servicer_url + post
            status, errorMessage, operation_time = postBody(url, value)
            if (status == 200 and errorMessage is None):
                sheet.cell(row=i, column=5).value = "Y"
                sheet.cell(row=i, column=6).value = status
                sheet.cell(row=i, column=7+number).value = operation_time
            elif (status == 200 and errorMessage is not None):
                sheet.cell(row=i, column=5).value = "N"
                sheet.cell(row=i, column=6).value = errorMessage
                sheet.cell(row=i, column=7+number).value = operation_time
            else:
                sheet.cell(row=i, column=5).value = "F"
                sheet.cell(row=i, column=6).value = status
                sheet.cell(row=i, column=7+number).value = operation_time
    ExcelMean(Excel)#求平均值
    ExcleTestResult(Excel)#填写测试结果
    TestPersonnel(Excel)#填写测试信息
    ExcleBorder(Excel)#设置边框
    Excel.save(Excel_File)#保存文档

def writeExcel(File):  #写入Excel内容
    FileTet =openpyxl.load_workbook(File)
    sheet =FileTet.active
    sheet.cell(row=2, column=5).value = 99
    sheet.cell(row=3, column=5, value=100)
    FileTet.save(File)


